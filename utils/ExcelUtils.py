from openpyxl import load_workbook
import os

base_dir = os.getcwd()
# TestData folder
testdata_dir = os.path.join(base_dir, "TestData")
# Create folder if not present
os.makedirs(testdata_dir, exist_ok=True)
# Excel file path
excel_file = os.path.join(testdata_dir, "testdata.xlsx")


def get_column_map(sheet):
    """
    Create a dictionary mapping column names -> column index
    """
    column_map = {}
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header:
            column_map[header.strip()] = col
    return column_map


def read_data():
    """
    Read data from excel file and mapping each column reference against field validation
    """

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    column_map = get_column_map(sheet)

    username_col = column_map["Username"]
    password_col = column_map["Password"]
    expected_test_column = column_map["NameofTest"]

    data = []

    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row=row, column=username_col).value
        password = sheet.cell(row=row, column=password_col).value
        expected = sheet.cell(row=row, column=expected_test_column).value
        data.append((row, username, password,expected))

    return data

